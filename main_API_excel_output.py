#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Mar 29 11:35:57 2022

@author: ryanschubert
"""
import requests
import pandas as pd
from datetime import date,timedelta
import numpy as np
import os
from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart,Reference,Series
import json
from math import isnan
#pull df of individuals
def identify_last_day_updates(apiToken,apiURL): #this function identifies each unique MRN that either has an updated response action or a created record action in the last 24 hours
    last_24_hours=str(date.today()-timedelta(days=1)) + " 00:00"
    data = {
        'token': apiToken,
        'content': 'log',
        'logtype': '',
        'user': '',
        'record': '',
        'beginTime': last_24_hours,
        'endTime': '',
        'format': 'json',
        'returnFormat': 'json'
    }
    r = requests.post(apiURL,data=data)
    r.raise_for_status()
    tmp=pd.DataFrame([json.loads(i) for i in r.json()])
    tmp=tmp[tmp['action'].str.contains("Update Response|Create Record")]
    mrn=tmp['action'].str.replace("Update Response|Create Record|(Auto calculation)","").str.replace("(","").str.replace(")","").str.strip().unique()
    return mrn

def extract_data(mrn,apiToken,apiURL): #this function extracts all exisiting pcl,ptci, and phq redcap data for a given mrn
    data = {
        'token': apiToken,
        'content': 'record',
        'action': 'export',
        'format': 'json',
        'type': 'flat',
        'csvDelimiter': '',
        'forms[0]': 'participant_information',
        'forms[1]': 'pcl_past_month',
        'forms[2]': 'phq9',
        'forms[3]': 'ptci',
        'forms[4]': 'pcl_past_week',
        'forms[5]': 'participant_validation',
        'records[0]': mrn,
        'events[0]': 'baseline_survey_arm_1',
        'events[1]': 'monday_week_1_arm_1',
        'events[2]': 'tuesday_week_1_arm_1',
        'events[3]': 'wednesday_week_1_arm_1',
        'events[4]': 'thursday_week_1_arm_1',
        'events[5]': 'friday_week_1_arm_1',
        'events[6]': 'monday_week_2_arm_1',
        'events[7]': 'tuesday_week_2_arm_1',
        'events[8]': 'wednesday_week_2_arm_1',
        'events[9]': 'thursday_week_2_arm_1',
        'events[10]': 'posttreatment_ques_arm_1',
        'events[11]': '1_month_followup_arm_1',
        'events[12]': '3_month_followup_arm_1',
        'events[13]': '6_month_followup_arm_1',
        'events[14]': '12_month_followup_arm_1',
        'events[15]': 'preiop_family_asse_arm_2',
        'events[16]': 'postiop_family_ass_arm_2',
        'events[17]': 'baseline_survey_arm_3',
        'events[18]': 'monday_week_1_arm_3',
        'events[19]': 'tuesday_week_1_arm_3',
        'events[20]': 'wednesday_week_1_arm_3',
        'events[21]': 'thursday_week_1_arm_3',
        'events[22]': 'friday_week_1_arm_3',
        'events[23]': 'monday_week_2_arm_3',
        'events[24]': 'tuesday_week_2_arm_3',
        'events[25]': 'wednesday_2_arm_3',
        'events[26]': 'thursday_week_2_arm_3',
        'events[27]': 'posttreatment_ques_arm_3',
        'events[28]': '3_month_followup_arm_3',
        'events[29]': '6_month_followup_arm_3',
        'events[30]': '12_month_followup_arm_3',
        'rawOrLabel': 'raw',
        'rawOrLabelHeaders': 'raw',
        'exportCheckboxLabel': 'false',
        'exportSurveyFields': 'false',
        'exportDataAccessGroups': 'false',
        'returnFormat': 'json'
    }
    
    r = requests.post(apiURL,data=data)
    r.raise_for_status()
    rD=pd.DataFrame(r.json())
    return rD

def extract_complete_record_set(mrns,apiToken,apiURL):
    complete_records=pd.DataFrame()
    for mrn in mrns:
        tmp=extract_data(mrn,apiToken,apiURL)
        complete_records=complete_records.append(tmp)
    complete_records.reset_index(drop=True,inplace=True)
    return complete_records

def clean_records(df):
    relevant_records=df[['mrn','first_name','last_name','startdate','date','redcap_event_name','cohort','pcl5_score','pcl5_score_pastmonth','phq9_score','phq9s_score','ptci_score']]
    relevant_records=relevant_records.assign(pcl5_score_all=np.where(relevant_records['pcl5_score']=="",relevant_records['pcl5_score_pastmonth'],relevant_records['pcl5_score']))
    relevant_records=relevant_records.loc[relevant_records['redcap_event_name'].str.contains('arm_1')]
    relevant_records=relevant_records.replace("",np.NaN)
    relevant_records['pcl5_score'] = relevant_records['pcl5_score'].astype(float)
    relevant_records['pcl5_score_pastmonth'] = relevant_records['pcl5_score_pastmonth'].astype(float)
    relevant_records['pcl5_score_all'] = relevant_records['pcl5_score_all'].astype(float)
    relevant_records['phq9_score'] = relevant_records['phq9_score'].astype(float)
    relevant_records['phq9s_score'] = relevant_records['phq9s_score'].astype(float)
    relevant_records['ptci_score'] = relevant_records['ptci_score'].astype(float)
    monovalue_records=relevant_records['mrn'].value_counts()[relevant_records['mrn'].value_counts()==1].index.tolist()
    relevant_records=relevant_records.loc[~relevant_records['mrn'].isin(monovalue_records)]
    condlist=[relevant_records['redcap_event_name'].str.contains("baseline"),
    relevant_records['redcap_event_name'].str.contains("monday_week_1_arm_1"),
    relevant_records['redcap_event_name'].str.contains("tuesday_week_1_arm_1"),
    relevant_records['redcap_event_name'].str.contains("wednesday_week_1_arm_1"),
    relevant_records['redcap_event_name'].str.contains("thursday_week_1_arm_1"),
    relevant_records['redcap_event_name'].str.contains("friday_week_1_arm_1"),
    relevant_records['redcap_event_name'].str.contains("monday_week_2_arm_1"),
    relevant_records['redcap_event_name'].str.contains("tuesday_week_2_arm_1"),
    relevant_records['redcap_event_name'].str.contains("wednesday_week_2_arm_1"),
    relevant_records['redcap_event_name'].str.contains("thursday_week_2_arm_1"),
    relevant_records['redcap_event_name'].str.contains("posttreatment_ques_arm_1"),
    relevant_records['redcap_event_name'].str.contains("1_month_followup_arm_1"),
    relevant_records['redcap_event_name'].str.contains("3_month_followup_arm_1"),
    relevant_records['redcap_event_name'].str.contains("6_month_followup_arm_1"),
    relevant_records['redcap_event_name'].str.contains("12_month_followup_arm_1")]
    choicelist=["Baseline",
                 "Mon w1",
                 "Tue w1",
                 "Wed w1",
                 "Thur w1",
                 "Fri w1",
                 "Mon w2",
                 "Tue w2",
                 "Wed w2",
                 "Thur w2",
                 "Post",
                 "1 Mon",
                 "3 Mon",
                 "6 Mon",
                 "12 Mon"]
    relevant_records['redcap_event_name']=np.select(condlist,choicelist,"null")
    return(relevant_records)


def subset_records(mrn,df):
    tmp=df.loc[df['mrn'] == mrn]
    return tmp

#extract the vector of cohort numbers
def lookup_cohort(df,mrn):
    cohort=df.loc[df['mrn']==mrn]['cohort']
    num=cohort[cohort.apply(lambda x: isinstance(x,str))]
    return num[num.index[0]]

def lookup_cohort_startdate(cohort_num,df):
    tmp=df.loc[(df['cohort']==cohort_num)]['mrn']
    tmp=df.loc[df['mrn'].isin(tmp)]
    tmp=tmp.loc[(tmp['redcap_event_name'] == "Mon w1")]['date']
    tmp=pd.to_datetime(tmp,format="%Y-%m-%d")
    sd=min(tmp).month_name() + " " + str(min(tmp).date().day)
    year=str(min(tmp).year)
    return sd, year

def lookup_initials(mrn,df):
     tmp=df.loc[df['mrn'] == mrn]
     tmp=tmp[['first_name','last_name']]
     tmp=tmp.loc[(tmp['first_name'] == tmp['first_name']) & (tmp['last_name'] == tmp['last_name'])]
     initials=tmp.iloc[0,0][0] + tmp.iloc[0,1][0]
     return(initials)

def get_outpath(root_out,sd,year): #lookup outdir or create one if none exists
    if not root_out.endswith("/"):
        root_out=root_out + "/"
    outpath=root_out + year + "/" + sd + "/"
    if not os.path.isdir(outpath):
        os.makedirs(outpath)
    return outpath

def initialize_workbook(outpath): #workbook initialization happens on a per cohort basis
    wb = Workbook()
    ws1=wb.active
    ws1.title = "Sheet 1"
    wb.save(outpath + "Score Tracking.xlsx")
    wb.close()
    return outpath + "Score Tracking.xlsx"


def target_workbook(outpath):
    files=os.listdir(outpath)
    if "Score Tracking.xlsx" in files:
        return(outpath + "Score Tracking.xlsx")
    else:
        file=initialize_workbook(outpath)
        return(file)

def initialize_sheet(targetWB,ID): #happens on a per individual basis
    wb = load_workbook(targetWB)
    newSheet=wb.create_sheet(ID)
    #add columns
    newSheet['A1'] = "Cohort"
    newSheet['B1'] = "MRN"
    newSheet['C1'] = "Date"
    newSheet['D1'] = "Day"
    newSheet['E1'] = "PCL-5"
    newSheet['F1'] = "PHQ-9"
    newSheet['G1'] = "PHQ-9s"
    newSheet['H1'] = "PTCI"
    
    newSheet['D2'] = "Baseline"
    newSheet['D3'] = "Mon w1"
    newSheet['D4'] = "Tue w1"
    newSheet['D5'] = "Wed w1"
    newSheet['D6'] = "Thur w1"
    newSheet['D7'] = "Fri w1"
    newSheet['D8'] = "Mon w2"
    newSheet['D9'] = "Tue w2"
    newSheet['D10'] = "Wed w2"
    newSheet['D11'] = "Thur w2"
    newSheet['D12'] = "Post"
    newSheet['D13'] = "1 Mon"
    newSheet['D14'] = "3 Mon"
    newSheet['D15'] = "6 Mon"
    newSheet['D16'] = "12 Mon"
    
    #add graphics
    c1 = LineChart()
    c1.title = "PCL-5"
    c1.style=13
    c1.legend=None
    c1.display_blanks = "span"
    c1.y_axis.scaling.min = 0
    c1.y_axis.scaling.max = 80
    data = Reference(newSheet,min_col=5,max_col=5,min_row=2,max_row=16)
    c1.series.append(Series(data))
    labels = Reference(newSheet,min_col=4,min_row=2,max_row=16)
    c1.set_categories(labels)
    s1=c1.series[0]
    s1.graphicalProperties.line.solidfill = "0066CC"
    s1.marker.symbol="triangle"
    s1.marker.graphicalProperties.solidFill="0066CC"
    newSheet.add_chart(c1, 'J2')
    
    c2 = LineChart()
    c2.title = "PHQ-9"
    c2.style=13
    c2.legend=None
    c2.display_blanks = "span"
    c2.y_axis.scaling.min = 0
    c2.y_axis.scaling.max = 27
    data = Reference(newSheet,min_col=6,min_row=2,max_row=16)
    c2.series.append(Series(data))
    labels = Reference(newSheet,min_col=4,min_row=2,max_row=16)
    c2.set_categories(labels)
    s2=c2.series[0]
    s2.graphicalProperties.line.solidfill = "0066CC"
    s2.marker.symbol="triangle"
    s2.marker.graphicalProperties.solidFill="0066CC"
    newSheet.add_chart(c2, 'S2')
    
    c3 = LineChart()
    c3.title = "PHQ-9s"
    c3.style=13
    c3.legend=None
    c3.display_blanks = "span"
    c3.y_axis.scaling.min = 0
    c3.y_axis.scaling.max = 3
    data = Reference(newSheet,min_col=7,min_row=2,max_row=16)
    c3.series.append(Series(data))
    labels = Reference(newSheet,min_col=4,min_row=2,max_row=16)
    c3.set_categories(labels)
    s3=c3.series[0]
    s3.graphicalProperties.line.solidfill = "0066CC"
    s3.marker.symbol="triangle"
    s3.marker.graphicalProperties.solidFill="0066CC"
    newSheet.add_chart(c3, 'J17')
    
    c4 = LineChart()
    c4.title = "PTCI"
    c4.style=13
    c4.legend=None
    c4.display_blanks = "span"
    c4.y_axis.scaling.min = 33
    c4.y_axis.scaling.max = 231
    data = Reference(newSheet,min_col=8,min_row=2,max_row=16)
    c4.series.append(Series(data))
    labels = Reference(newSheet,min_col=4,min_row=2,max_row=16)
    c4.set_categories(labels)
    s4=c4.series[0]
    s4.graphicalProperties.line.solidfill = "0066CC"
    s4.marker.symbol="triangle"
    s4.marker.graphicalProperties.solidFill="0066CC"
    newSheet.add_chart(c4, 'S17')
    wb.save(targetWB)
    wb.close()
    

def update_sheet(df,targetWB,ID):
    condlist=[df['redcap_event_name'].str.contains("Baseline"),
       df['redcap_event_name'].str.contains("Mon w1"),
       df['redcap_event_name'].str.contains("Tue w1"),
       df['redcap_event_name'].str.contains("Wed w1"),
       df['redcap_event_name'].str.contains("Thur w1"),
       df['redcap_event_name'].str.contains("Fri w1"),
       df['redcap_event_name'].str.contains("Mon w2"),
       df['redcap_event_name'].str.contains("Tue w2"),
       df['redcap_event_name'].str.contains("Wed w2"),
       df['redcap_event_name'].str.contains("Thur w2"),
       df['redcap_event_name'].str.contains("Post"),
       df['redcap_event_name'].str.contains("1 Mon"),
       df['redcap_event_name'].str.contains("3 Mon"),
       df['redcap_event_name'].str.contains("6 Mon"),
       df['redcap_event_name'].str.contains("12 Mon")]
    choicelist=["2",
                "3",
                "4",
                "5",
                "6",
                "7",
                "8",
                "9",
                "10",
                "11",
                "12",
                "13",
                "14",
                "15",
                "16"]
    df['row']=np.select(condlist,choicelist,"null")
    df=df.loc[df.row != "null"]
    
    wb = load_workbook(targetWB)
    if ID not in wb.sheetnames:
        wb.close()
        initialize_sheet(targetWB,ID)
        wb = load_workbook(targetWB)
    ws = wb[ID]
    for index, row in df.iterrows():
        r=row['row']
        ws['A' + r]=row['cohort']
        ws['B' + r]=row['mrn']
        ws['C' + r]=row['date']
        ws['E' + r]=row['pcl5_score_all']
        ws['F' + r]=row['phq9_score']
        ws['G' + r]=row['phq9s_score']
        ws['H' + r]=row['ptci_score']
    wb.save(targetWB)
    wb.close()
    
    


def preverify_SSL():
    try:
        session = requests.Session()
        response = session.get('https://redcap.rush.edu/redcap/',allow_redirects=False,verify=True)
    except requests.exceptions.SSLError:
        print("Error: problems verifying SSL ")
        print(requests.exceptions.SSLError)
        exit()


def main(apiToken,apiURL,root_out):
    preverify_SSL()
    mrns=identify_last_day_updates(apiToken,apiURL)
    records=extract_complete_record_set(mrns=mrns,apiToken=apiToken,apiURL=apiURL)
    records=clean_records(df=records)
    mrns=records.mrn.unique()
    for mrn in mrns:
        print(mrn)
        cohort_number=lookup_cohort(mrn=mrn,df=records)
        start,year=lookup_cohort_startdate(cohort_num=cohort_number,df=records)
        initials=lookup_initials(mrn=mrn,df=records)
        outpath=get_outpath(root_out=root_out,sd=start,year=year)
        targetWB=target_workbook(outpath)
        subset=subset_records(mrn=mrn, df=records)
        update_sheet(df=subset,targetWB=targetWB, ID = initials + "_" + mrn)
        
